"""
一次性生成 dice.xlsx —— 骰子表
对应 common/autoload/game_data.gd 的 _register_xxx_dice

三表结构：
- base 25 行：骰子身份 + effect_group 外键
- effects 多行：每个 effect_group 对应多条效果
- status 少量行：复杂状态组合（当前用 ST 紧凑编码，这张表保留空架子）

运行：python config/tools/build_excel_dice.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
EXCEL_DIR = os.path.join(PROJECT_ROOT, "config", "excel")

HEADER_FILL_NAME = PatternFill(start_color="FFB6D7A8", end_color="FFB6D7A8", fill_type="solid")
HEADER_FILL_TYPE = PatternFill(start_color="FFFFE599", end_color="FFFFE599", fill_type="solid")
HEADER_FILL_DESC = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")


def write_header(ws, headers, widths=None):
    for col, (name, type_, desc) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = HEADER_FILL_NAME
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=2, column=col, value=type_)
        c.fill = HEADER_FILL_TYPE
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=3, column=col, value=desc)
        c.fill = HEADER_FILL_DESC
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A4"
    for col in range(1, len(headers) + 1):
        w = widths[col - 1] if widths else 16
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 70


def build_dice():
    wb = Workbook()

    # ===== Sheet 1: base =====
    ws = wb.active
    ws.title = "base"
    write_header(ws, [
        ("id", "string", "骰子编号 D+4位。D0001-0099通用/D0101+战士/D0201+法师/D0301+盗贼"),
        ("legacy_key", "string", "代码旧 key（standard/w_bloodthirst 等）迁移参照"),
        ("class", "ref", "所属职业 C+2位。C00=通用"),
        ("name", "string", "中文名"),
        ("faces", "intarray", "6个点数面，逗号分隔。例 1,2,3,4,5,6；诅咒骰 0,0,0,0,0,0"),
        ("rarity", "enum", "稀有度 RA1-5。见 _enums.xlsx"),
        ("element", "enum", "元素 EL0-6。EL0 无元素，EL1 火..."),
        ("description", "string", "卡面描述"),
        ("effect_group", "ref", "效果组外键 EG+4位。空=纯点数骰无效果"),
        ("flag_cursed", "bool", "诅咒骰标记（商店剔除）1是0否"),
        ("flag_cracked", "bool", "碎裂骰标记（出牌后自残）1是0否"),
    ], widths=[8, 18, 6, 12, 16, 6, 6, 32, 12, 10, 10])

    # 格式：(id, legacy, class, name, faces, rarity, element, desc, eg, cursed, cracked)
    # rarity: RA1=COMMON, RA2=UNCOMMON, RA3=RARE, RA4=LEGENDARY, RA5=CURSE
    # (代码里 RARE 对应 RA3，game_data 里用的是 DiceRarity.RARE)
    base_rows = [
        # 通用
        ("D0001", "standard", "C00", "普通骰子", "1,2,3,4,5,6", "RA1", "EL0", "标准六面骰", "", 0, 0),
        ("D0002", "blade", "C00", "锋刃骰子", "1,2,3,4,5,6", "RA3", "EL0", "出牌时追加5点固定伤害", "EG0002", 0, 0),
        ("D0003", "amplify", "C00", "倍增骰子", "1,2,3,4,5,6", "RA3", "EL0", "出牌时最终伤害提升20%", "EG0003", 0, 0),
        ("D0004", "split", "C00", "分裂骰子", "1,2,3,4,5,6", "RA3", "EL0", "出牌时分裂出1颗相同点数的临时骰子", "EG0004", 0, 0),
        ("D0005", "magnet", "C00", "磁吸骰子", "1,2,3,4,5,6", "RA3", "EL0", "出牌时随机将1颗同伴骰子点数变为与本骰子相同", "EG0005", 0, 0),
        ("D0006", "joker", "C00", "小丑骰子", "1,2,3,4,5,6,7,8,9", "RA3", "EL0", "点数1到9随机，突破六面骰限制", "EG0006", 0, 0),
        ("D0007", "chaos", "C00", "混沌骰子", "1,1,1,6,6,6", "RA4", "EL0", "只会掷出1或6", "EG0007", 0, 0),
        ("D0008", "cursed", "C00", "诅咒骰子", "0,0,0,0,0,0", "RA5", "EL0", "点数固定0", "", 1, 0),
        ("D0009", "cracked", "C00", "碎裂骰子", "1,1,1,2,2,2", "RA5", "EL0", "出牌后受2点反噬", "EG0009", 0, 1),
        ("D0010", "temp_rogue", "C00", "暗影残骰", "1,1,2,2,3,3", "RA1", "EL0", "连击奖励临时骰子", "", 0, 0),
        ("D0011", "heavy", "C00", "灌铅骰子", "4,4,5,5,6,6", "RA2", "EL0", "只会掷出4/5/6", "", 0, 0),
        # 战士
        ("D0101", "w_bloodthirst", "C01", "嗜血骰子", "1,2,3,4,5,6", "RA2", "EL0", "卖血重投时+3伤害", "EG0101", 0, 0),
        ("D0102", "w_ironwall", "C01", "铁壁骰子", "1,2,3,4,5,6", "RA2", "EL0", "出牌时获得等同点数的护甲", "EG0102", 0, 0),
        ("D0103", "w_fury", "C01", "怒火骰子", "1,2,3,4,5,6", "RA3", "EL0", "受到敌人攻击时永久+1伤害", "EG0103", 0, 0),
        ("D0104", "w_execute", "C01", "处刑骰子", "1,2,3,4,5,6", "RA3", "EL0", "敌人HP≤30%时伤害翻倍", "EG0104", 0, 0),
        ("D0105", "w_berserker", "C01", "狂暴骰子", "1,2,3,4,5,6", "RA3", "EL0", "自残5点HP，伤害+50%", "EG0105", 0, 0),
        # 法师
        ("D0201", "mage_elemental", "C02", "元素骰子", "1,2,3,4,5,6", "RA2", "EL0", "每回合随机变为火/冰/雷/毒/圣元素", "EG0201", 0, 0),
        ("D0202", "mage_reverse", "C02", "逆转骰子", "1,2,3,4,5,6", "RA2", "EL0", "出牌时点数翻转(7-点数)", "EG0202", 0, 0),
        ("D0203", "mage_crystal", "C02", "水晶骰子", "1,2,3,4,5,6", "RA3", "EL0", "保留到下回合时点数+1", "EG0203", 0, 0),
        ("D0204", "mage_stardust", "C02", "星尘骰子", "1,2,3,4,5,6", "RA3", "EL0", "每保留1回合+1点(上限5)", "EG0204", 0, 0),
        # 盗贼
        ("D0301", "r_quickdraw", "C03", "快攻骰子", "1,2,3,4,5,6", "RA2", "EL0", "连击时+20%伤害", "EG0301", 0, 0),
        ("D0302", "r_combomastery", "C03", "连击心得", "1,2,3,4,5,6", "RA2", "EL0", "连击时获得暗影残骰", "EG0302", 0, 0),
        ("D0303", "r_poisondart", "C03", "毒镖骰子", "1,2,3,4,5,6", "RA3", "EL0", "附加2层中毒", "EG0303", 0, 0),
        ("D0304", "r_shadowclone", "C03", "影分身", "1,2,3,4,5,6", "RA3", "EL0", "出牌时复制自身一同结算", "EG0304", 0, 0),
    ]
    for r, data in enumerate(base_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws.cell(row=r, column=col, value=v)

    # ===== Sheet 2: effects =====
    ws_ef = wb.create_sheet("effects")
    write_header(ws_ef, [
        ("effect_group", "ref", "效果组编号 EG+4位。多行同组=该组含多条效果"),
        ("effect_type", "enum", "效果类型 ET+2位。见 _enums.xlsx"),
        ("param_key", "string", "参数键名（对应代码字段，bonus_damage/bonus_mult/armor 等）"),
        ("param_value", "string", "参数值。数字/布尔/字符串。布尔用 1/0"),
        ("param_ref", "string", "引用型参数。状态：ST+类型x数值x持续，如 ST01x2x3"),
        ("note", "string", "策划备注（不进代码）"),
    ], widths=[14, 12, 22, 12, 18, 28])

    # 每个 effect_group 的效果条目
    # 字段名与 DiceDef 的 @export 名严格对应
    effects_rows = [
        # EG0002 锋刃 bonus_damage 5
        ("EG0002", "ET01", "bonus_damage", 5, "", "锋刃 +5 固定伤害"),
        # EG0003 倍增 bonus_mult 1.2
        ("EG0003", "ET02", "bonus_mult", 1.2, "", "倍增 ×1.2"),
        # EG0004 分裂（仅描述，代码里靠其他机制；这里不写具体字段，注释占位）
        ("EG0004", "ET12", "split_dice", 1, "", "分裂骰子（特殊逻辑，代码里判断）"),
        # EG0005 磁吸
        ("EG0005", "ET12", "magnet_dice", 1, "", "磁吸骰子（特殊逻辑）"),
        # EG0006 小丑（无实际字段，faces 不同而已）
        ("EG0006", "ET12", "joker_dice", 1, "", "小丑骰子 faces 特殊"),
        # EG0007 混沌（faces 不同）
        ("EG0007", "ET12", "chaos_dice", 1, "", "混沌骰 faces 特殊"),
        # EG0009 碎裂 self_damage 2
        ("EG0009", "ET06", "self_damage", 2, "", "碎裂 -2 HP"),
        # EG0101 嗜血 scale_with_blood_rerolls + bonus_damage 3
        ("EG0101", "ET11", "scale_with_blood_rerolls", 1, "", "嗜血重投时生效"),
        ("EG0101", "ET01", "bonus_damage", 3, "", "嗜血 +3 伤害"),
        # EG0102 铁壁 armor_from_value
        ("EG0102", "ET11", "armor_from_value", 1, "", "护甲 = 点数"),
        # EG0103 怒火（被动无字段，由外部逻辑响应 on_damage_taken）
        ("EG0103", "ET12", "fury_stack", 1, "", "受击永久+1伤害（代码写死）"),
        # EG0104 处刑 execute_threshold 0.3 + execute_mult 2.0
        ("EG0104", "ET11", "execute_threshold", 0.3, "", "敌人 HP <= 30% 触发"),
        ("EG0104", "ET02", "execute_mult", 2.0, "", "处刑 ×2"),
        # EG0105 狂暴 self_damage 5 + self_berserk + bonus_mult 1.5
        ("EG0105", "ET06", "self_damage", 5, "", "狂暴自残 5"),
        ("EG0105", "ET11", "self_berserk", 1, "", "狂暴标记"),
        ("EG0105", "ET02", "bonus_mult", 1.5, "", "狂暴 ×1.5"),
        # EG0201 元素骰 is_elemental
        ("EG0201", "ET11", "is_elemental", 1, "", "每回合随机元素"),
        # EG0202 逆转 reverse_value
        ("EG0202", "ET11", "reverse_value", 1, "", "点数 = 7 - 原点数"),
        # EG0203 水晶 bonus_on_keep 1
        ("EG0203", "ET10", "bonus_on_keep", 1, "", "保留 +1 点"),
        # EG0204 星尘 bonus_per_turn_kept 1 + keep_bonus_cap 5
        ("EG0204", "ET10", "bonus_per_turn_kept", 1, "", "每保留1回合 +1"),
        ("EG0204", "ET10", "keep_bonus_cap", 5, "", "保留加成上限 5"),
        # EG0301 快攻 combo_bonus 0.2
        ("EG0301", "ET11", "combo_bonus", 0.2, "", "连击 +20%"),
        # EG0302 连击心得 grant_shadow_die
        ("EG0302", "ET09", "grant_shadow_die", 1, "", "连击获得暗影残骰"),
        # EG0303 毒镖 poison_base 2 + status_to_enemy
        ("EG0303", "ET11", "poison_base", 2, "", "毒镖基础毒值"),
        ("EG0303", "ET07", "status_to_enemy", "", "ST01x2x3", "施加 2 层中毒（类型ST01/POISON）持续3回合"),
        # EG0304 影分身 shadow_clone_play
        ("EG0304", "ET11", "shadow_clone_play", 1, "", "复制自身结算"),
    ]
    for r, data in enumerate(effects_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws_ef.cell(row=r, column=col, value=v)

    # ===== Sheet 3: status（占位）=====
    ws_st = wb.create_sheet("status")
    write_header(ws_st, [
        ("status_id", "string", "状态组自定义 id，仅在 param_ref 紧凑编码无法表达时使用"),
        ("target", "string", "enemy 或 self"),
        ("status_type", "enum", "状态类型 ST+2位"),
        ("value", "int", "数值"),
        ("duration", "int", "持续回合"),
        ("note", "string", "备注"),
    ], widths=[14, 10, 10, 8, 10, 24])
    # 目前无复杂组合，仅表头

    out_path = os.path.join(EXCEL_DIR, "dice.xlsx")
    wb.save(out_path)
    print(f"[OK] {out_path} · base={len(base_rows)} effects={len(effects_rows)}")


if __name__ == "__main__":
    build_dice()
