"""
一次性生成 class.xlsx —— 职业表
对应 data/class_def.gd 的 3 个 static func

Sheet 1: base         职业身份与属性
Sheet 2: starting_items  起始骰子/遗物（一对多）
Sheet 3: skills       技能描述

运行：python config/tools/build_excel_class.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
EXCEL_DIR = os.path.join(PROJECT_ROOT, "config", "excel")

HEADER_FILL_NAME = PatternFill(start_color="FFB6D7A8", end_color="FFB6D7A8", fill_type="solid")
HEADER_FILL_TYPE = PatternFill(start_color="FFFFE599", end_color="FFFFE599", fill_type="solid")
HEADER_FILL_DESC = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")


def write_header(ws, headers, widths=None):
    for col, (name, type_, desc) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = HEADER_FILL_NAME
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=2, column=col, value=type_)
        c.fill = HEADER_FILL_TYPE
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=3, column=col, value=desc)
        c.fill = HEADER_FILL_DESC
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A4"
    for col in range(1, len(headers) + 1):
        w = widths[col - 1] if widths else 20
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 60


# 旧 id → 新编号 映射表（后续 ID 映射表会读这个）
# 职业
CLASS_ID_MAP = {
    "warrior": "C01",
    "mage": "C02",
    "rogue": "C03",
}

# 骰子（完整映射，class.xlsx 的 starting_items 会引用；dice.xlsx 生成时也用同一份）
DICE_ID_MAP = {
    # 通用
    "standard": "D0001",
    "blade": "D0002",
    "amplify": "D0003",
    "split": "D0004",
    "magnet": "D0005",
    "joker": "D0006",
    "chaos": "D0007",
    "cursed": "D0008",
    "cracked": "D0009",
    "temp_rogue": "D0010",
    "heavy": "D0011",
    # 战士
    "w_bloodthirst": "D0101",
    "w_ironwall": "D0102",
    "w_fury": "D0103",
    "w_execute": "D0104",
    "w_berserker": "D0105",
    # 法师
    "mage_elemental": "D0201",
    "mage_reverse": "D0202",
    "mage_crystal": "D0203",
    "mage_stardust": "D0204",
    # 盗贼
    "r_quickdraw": "D0301",
    "r_combomastery": "D0302",
    "r_poisondart": "D0303",
    "r_shadowclone": "D0304",
}


def dice_ids_from_legacy(legacy_list):
    """把旧 id 列表转成分号分隔的新编号字符串"""
    return ";".join(DICE_ID_MAP[x] for x in legacy_list)


def build_class():
    wb = Workbook()

    # ===== Sheet 1: base =====
    ws = wb.active
    ws.title = "base"
    write_header(ws, [
        ("id", "string", "职业编号 C+2位（C01/C02/C03）"),
        ("legacy_key", "string", "代码旧 key（warrior/mage/rogue）迁移参照"),
        ("name", "string", "职业名"),
        ("title", "string", "称号"),
        ("description", "string", "职业描述"),
        ("color", "string", "主色 #RRGGBB"),
        ("color_light", "string", "亮色 #RRGGBB"),
        ("color_dark", "string", "暗色 #RRGGBB"),
        ("hp", "int", "初始 HP（同步为 max_hp）"),
        ("draw_count", "int", "每回合抽骰数"),
        ("max_plays", "int", "每回合出牌次数"),
        ("free_rerolls", "int", "每回合免费重投"),
        ("flag_blood_reroll", "bool", "是否可嗜血重投（战士专属），1是0否"),
        ("flag_keep_unplayed", "bool", "是否保留未出牌骰（法师专属），1是0否"),
        ("flag_multi_select", "bool", "普攻是否可多选（战士专属），1是0否"),
        ("passive_desc", "string", "被动技能总描述"),
    ], widths=[8, 12, 14, 14, 40, 12, 12, 12, 6, 10, 10, 10, 14, 14, 14, 50])

    warriors = [
        "C01", "warrior", "嗜血狂战", "铁血征服者",
        "以鲜血为代价，换取毁天灭地的一击。嗜血越多，伤害越高。",
        "#c04040", "#ff6060", "#601010",
        120, 3, 1, 1, 1, 0, 1,
        "【血怒战意】嗜血每次+15%最终伤害（最多5层+75%）；叠满后卖血改为+5护甲；血量≤50%手牌+1；手牌溢出上限时按受伤百分比加伤害倍率；普攻可多选",
    ]
    mages = [
        "C02", "mage", "星界魔导", "星界禁咒师",
        "耐心吟唱，两三回合攒齐完美手牌，打出毁天灭地的大招。",
        "#7040c0", "#a070ff", "#301060",
        100, 3, 1, 1, 0, 1, 0,
        "【星界吟唱】未出牌骰子保留到下回合（3→4→5→6递增）；吟唱回合获得递增护甲；满6后继续吟唱每次+10%伤害；出牌后重置",
    ]
    rogues = [
        "C03", "rogue", "影锋刺客", "暗影连击者",
        "一回合出牌两次，连击加成层层递增。暗影残骰是连击的灵魂。",
        "#30a050", "#60d080", "#104020",
        90, 3, 2, 1, 0, 0, 0,
        "【连击】每回合出牌2次；第2次伤害+20%；同牌型再+25%；暗影残骰是连击核心",
    ]
    for r, data in enumerate([warriors, mages, rogues], start=4):
        for col, v in enumerate(data, start=1):
            ws.cell(row=r, column=col, value=v)

    # ===== Sheet 2: starting_items =====
    ws_si = wb.create_sheet("starting_items")
    write_header(ws_si, [
        ("class_id", "ref", "职业编号 C+2位，指向 base sheet"),
        ("item_type", "string", "物品类型：dice 或 relic"),
        ("item_id", "ref", "物品编号（D+4位 或 R+4位）"),
        ("legacy_key", "string", "旧 key，调试用"),
        ("note", "string", "备注"),
    ], widths=[10, 10, 10, 18, 24])

    # 战士
    warrior_dice = ["standard", "standard", "standard", "standard", "w_bloodthirst", "w_ironwall"]
    mage_dice = ["standard", "standard", "standard", "standard", "mage_elemental", "mage_reverse"]
    rogue_dice = ["standard", "standard", "standard", "r_quickdraw", "r_combomastery"]

    start_items_rows = []
    for d in warrior_dice:
        start_items_rows.append(("C01", "dice", DICE_ID_MAP[d], d, ""))
    for d in mage_dice:
        start_items_rows.append(("C02", "dice", DICE_ID_MAP[d], d, ""))
    for d in rogue_dice:
        start_items_rows.append(("C03", "dice", DICE_ID_MAP[d], d, ""))

    for r, data in enumerate(start_items_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws_si.cell(row=r, column=col, value=v)

    # ===== Sheet 3: skills =====
    ws_sk = wb.create_sheet("skills")
    write_header(ws_sk, [
        ("class_id", "ref", "职业编号"),
        ("skill_idx", "int", "技能序号 0起"),
        ("name", "string", "技能名"),
        ("description", "string", "技能描述"),
    ], widths=[10, 10, 14, 60])

    skills = [
        ("C01", 0, "血怒战意", "每次嗜血，最终伤害+15%（最多叠加5层+75%），叠满后卖血获得5点护甲"),
        ("C01", 1, "狂暴本能", "血量≤50%时手牌上限+1颗；手牌达到6颗上限时，按受伤百分比获得等比例伤害倍率加成"),
        ("C01", 2, "铁拳连打", "普攻牌型可多选骰子，一次打出所有选中骰子的伤害"),
        ("C02", 0, "星界吟唱", "未出牌的骰子保留到下回合，手牌上限逐层递增（3→4→5→6）"),
        ("C02", 1, "吟唱护盾", "每次吟唱（不出牌）获得递增护甲（6/8/10/12...），层数越高护甲越厚"),
        ("C02", 2, "过充释放", "手牌满6颗后继续吟唱，每回合额外+10%伤害倍率；出牌后重置"),
        ("C03", 0, "双刃连击", "每回合可出牌2次，第2次出牌伤害+20%"),
        ("C03", 1, "精准连击", "两次出牌使用相同牌型时（非普攻），额外+25%伤害加成"),
        ("C03", 2, "暗影残骰", "连击触发时补充暗影残骰；连击奖励的残骰可保留到下回合"),
    ]
    for r, data in enumerate(skills, start=4):
        for col, v in enumerate(data, start=1):
            ws_sk.cell(row=r, column=col, value=v)

    out_path = os.path.join(EXCEL_DIR, "class.xlsx")
    wb.save(out_path)
    print(f"[OK] {out_path}")


if __name__ == "__main__":
    build_class()
