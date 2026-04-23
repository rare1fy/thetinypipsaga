"""
一次性生成 relic.xlsx —— 遗物表
对应 common/autoload/game_data.gd 的 _register_relics

两表结构：
- base 15 行：遗物身份 + trigger + effect_group 外键
- effects 多行：每个 effect_group 的参数

运行：python config/tools/build_excel_relic.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
EXCEL_DIR = os.path.join(PROJECT_ROOT, "config", "excel")

HEADER_FILL_NAME = PatternFill(start_color="FFB6D7A8", end_color="FFB6D7A8", fill_type="solid")
HEADER_FILL_TYPE = PatternFill(start_color="FFFFE599", end_color="FFFFE599", fill_type="solid")
HEADER_FILL_DESC = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")


def write_header(ws, headers, widths=None):
    for col, (name, type_, desc) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = HEADER_FILL_NAME
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=2, column=col, value=type_)
        c.fill = HEADER_FILL_TYPE
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=3, column=col, value=desc)
        c.fill = HEADER_FILL_DESC
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A4"
    for col in range(1, len(headers) + 1):
        w = widths[col - 1] if widths else 16
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 60


# 旧 id → 新编号
RELIC_ID_MAP = {
    "iron_heart": "R0001",
    "healing_herb": "R0002",
    "sharp_blade": "R0003",
    "lucky_coin": "R0004",
    "hourglass": "R0005",
    "fortune_wheel_relic": "R0006",
    "blood_pact": "R0007",
    "magic_glove": "R0008",
    "whetstone": "R0009",
    "rage_fire": "R0010",
    "prism_focus": "R0011",
    "limit_breaker": "R0012",
    "pair_upgrade": "R0013",
    "straight_master": "R0014",
    "soul_crystal": "R0015",
    "life_furnace": "R0016",
}


def build_relic():
    wb = Workbook()
    ws = wb.active
    ws.title = "base"

    write_header(ws, [
        ("id", "string", "遗物编号 R+4位"),
        ("legacy_key", "string", "代码旧 key（iron_heart 等）"),
        ("name", "string", "中文名"),
        ("rarity", "enum", "稀有度 RA1-4（遗物无诅咒级）"),
        ("trigger", "enum", "触发时机 TR+2位。见 _enums.xlsx"),
        ("description", "string", "遗物描述"),
        ("effect_group", "ref", "效果组 EG+4位（1001起为遗物段）"),
        ("flag_consume", "bool", "消耗型遗物（触发后消失）1是0否"),
        ("flag_unique", "bool", "唯一（一次只能持有一个）1是0否"),
    ], widths=[8, 20, 14, 8, 10, 44, 12, 10, 10])

    # rarity: COMMON=RA1, UNCOMMON=RA2, RARE=RA3, LEGENDARY=RA4
    # trigger: ON_PLAY=TR01, ON_KILL=TR02, ON_REROLL=TR03, ON_TURN_START=TR04, ON_TURN_END=TR05,
    #          ON_BATTLE_START=TR06, ON_BATTLE_END=TR07, ON_DAMAGE_TAKEN=TR08, ON_FATAL=TR09,
    #          ON_FLOOR_CLEAR=TR10, ON_MOVE=TR11, PASSIVE=TR99
    base_rows = [
        # (id, legacy, name, rarity, trigger, desc, eg, consume, unique)
        ("R0001", "iron_heart", "铁之心", "RA1", "TR06", "获得5点护甲", "EG1001", 0, 0),
        ("R0002", "healing_herb", "治愈草药", "RA1", "TR06", "战斗开始时恢复10HP", "EG1002", 0, 0),
        ("R0003", "sharp_blade", "锋利之刃", "RA1", "TR01", "每次出牌+3伤害", "EG1003", 0, 0),
        ("R0004", "lucky_coin", "幸运金币", "RA1", "TR99", "金币获取+20%", "EG1004", 0, 0),
        ("R0005", "hourglass", "时光沙漏", "RA3", "TR09", "致命伤害时免死一次，消耗此遗物", "EG1005", 1, 1),
        ("R0006", "fortune_wheel_relic", "命运之轮", "RA2", "TR99", "首次出牌后保留手牌一次", "EG1006", 0, 0),
        ("R0007", "blood_pact", "血之契约", "RA3", "TR05", "回合结束保留1颗最高点骰子", "EG1007", 0, 0),
        ("R0008", "magic_glove", "魔法手套", "RA2", "TR05", "每场战斗下回合+1手牌", "EG1008", 0, 0),
        ("R0009", "whetstone", "磨砺石", "RA3", "TR05", "每场战斗下回合+1出牌", "EG1009", 0, 0),
        ("R0010", "rage_fire", "怒火燎原", "RA2", "TR08", "受到伤害后下次出牌+5伤害", "EG1010", 0, 0),
        ("R0011", "prism_focus", "棱镜聚焦", "RA3", "TR99", "锁定一个元素，同元素牌型伤害+30%", "EG1011", 0, 0),
        ("R0012", "limit_breaker", "突破极限", "RA4", "TR99", "小丑骰子可掷出10-12", "EG1012", 0, 0),
        ("R0013", "pair_upgrade", "对子大师", "RA3", "TR99", "对子视为三条", "EG1013", 0, 0),
        ("R0014", "straight_master", "顺子大师", "RA3", "TR99", "顺子牌型等级+1", "EG1014", 0, 0),
        ("R0015", "soul_crystal", "魂晶之心", "RA4", "TR02", "溢出伤害×倍率×15%转化为魂晶", "EG1015", 0, 0),
        ("R0016", "life_furnace", "生命熔炉", "RA4", "TR01", "每出牌5次恢复15HP", "EG1016", 0, 0),
    ]
    for r, data in enumerate(base_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws.cell(row=r, column=col, value=v)

    # ===== Sheet 2: effects =====
    ws_ef = wb.create_sheet("effects")
    write_header(ws_ef, [
        ("effect_group", "ref", "遗物效果组 EG+4位"),
        ("effect_type", "enum", "效果类型 ET+2位"),
        ("param_key", "string", "参数键（armor/heal/damage/multiplier 等，与 RelicDef 字段对应）"),
        ("param_value", "string", "参数值"),
        ("note", "string", "备注"),
    ], widths=[14, 12, 26, 14, 28])

    effects_rows = [
        # EG1001 铁之心 armor 5
        ("EG1001", "ET03", "armor", 5, "铁之心 +5 护甲"),
        # EG1002 治愈草药 heal 10
        ("EG1002", "ET04", "heal", 10, "草药 回 10 HP"),
        # EG1003 锋利之刃 damage 3
        ("EG1003", "ET01", "damage", 3, "锋利 +3 伤害"),
        # EG1004 幸运金币 gold_bonus 20
        ("EG1004", "ET20", "gold_bonus", 20, "+20% 金币"),
        # EG1005 时光沙漏 prevent_death
        ("EG1005", "ET12", "prevent_death", 1, "免死一次"),
        # EG1006 命运之轮 keep_unplayed_once
        ("EG1006", "ET12", "keep_unplayed_once", 1, "保留未出牌一次"),
        # EG1007 血契 keep_highest_die 1
        ("EG1007", "ET12", "keep_highest_die", 1, "保留最高点 1 颗"),
        # EG1008 魔法手套 temp_draw_bonus 1
        ("EG1008", "ET12", "temp_draw_bonus", 1, "+1 手牌（下回合）"),
        # EG1009 磨砺石 grant_extra_play 1
        ("EG1009", "ET12", "grant_extra_play", 1, "+1 出牌（下回合）"),
        # EG1010 怒火燎原 damage 5
        ("EG1010", "ET01", "damage", 5, "受伤后 +5 伤害"),
        # EG1011 棱镜聚焦 multiplier 0.3
        ("EG1011", "ET02", "multiplier", 0.3, "同元素 +30%"),
        # EG1012 突破极限 max_points_unlocked
        ("EG1012", "ET12", "max_points_unlocked", 1, "小丑骰突破上限"),
        # EG1013 对子大师 pair_as_triplet
        ("EG1013", "ET12", "pair_as_triplet", 1, "对子视为三条"),
        # EG1014 顺子大师 straight_upgrade 1
        ("EG1014", "ET12", "straight_upgrade", 1, "顺子等级+1"),
        # EG1015 魂晶之心 multiplier 0.15
        ("EG1015", "ET02", "multiplier", 0.15, "溢伤 ×15% 转魂晶"),
        # EG1016 生命熔炉 heal 15 + counter
        ("EG1016", "ET04", "heal", 15, "每 5 出牌回 15 HP"),
        ("EG1016", "ET12", "max_counter", 5, "计数上限"),
    ]
    for r, data in enumerate(effects_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws_ef.cell(row=r, column=col, value=v)

    out_path = os.path.join(EXCEL_DIR, "relic.xlsx")
    wb.save(out_path)
    print(f"[OK] {out_path} · base={len(base_rows)} effects={len(effects_rows)}")


if __name__ == "__main__":
    build_relic()
