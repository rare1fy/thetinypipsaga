"""
一次性生成 enemy.xlsx —— 敌人表（4 sheet）
对应 data/enemy_config.gd 的 _build_all_configs

- base 38 行：敌人身份 + 分类 + 章节 + HP/DMG + 掉落 + phase_group/quote_group 外键
- phases 多行：每个 phase_group 下的阶段定义（HP 阈值）
- actions 多行：每个 phase 下的行动（ATTACK/DEFEND/SKILL）
- quotes 多行：每个 quote_group 的台词（入场/死亡/攻击/受击/低血）

运行：python config/tools/build_excel_enemy.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
EXCEL_DIR = os.path.join(PROJECT_ROOT, "config", "excel")

HEADER_FILL_NAME = PatternFill(start_color="FFB6D7A8", end_color="FFB6D7A8", fill_type="solid")
HEADER_FILL_TYPE = PatternFill(start_color="FFFFE599", end_color="FFFFE599", fill_type="solid")
HEADER_FILL_DESC = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")


def write_header(ws, headers, widths=None):
    for col, (name, type_, desc) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = HEADER_FILL_NAME
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=2, column=col, value=type_)
        c.fill = HEADER_FILL_TYPE
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=3, column=col, value=desc)
        c.fill = HEADER_FILL_DESC
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A4"
    for col in range(1, len(headers) + 1):
        w = widths[col - 1] if widths else 14
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 60


# ==== 章节段号约定 ====
# 章1 森林 E0001- / 章2 冰封 E0101- / 章3 熔岩 E0201- / 章4 暗影 E0301- / 章5 永恒 E0401-
# 精英 E9001-9099 / Boss E9101-9199
# 战斗类型：WARRIOR=1 / GUARDIAN=2 / RANGER=3 / CASTER=4 / PRIEST=5 → 用代码里枚举名
# category: NORMAL / ELITE / BOSS

# 敌人 legacy_id → 新编号
ENEMY_ID_MAP = {
    # 章1
    "forest_ghoul": "E0001",
    "forest_spider": "E0002",
    "forest_treant": "E0003",
    "forest_banshee": "E0004",
    "forest_wolf_priest": "E0005",
    # 章2
    "ice_yeti": "E0101",
    "ice_mage": "E0102",
    "ice_wolf": "E0103",
    "ice_golem": "E0104",
    # 章3
    "lava_hound": "E0201",
    "lava_imp": "E0202",
    "lava_guardian": "E0203",
    "lava_shaman": "E0204",
    # 章4
    "shadow_assassin": "E0301",
    "shadow_felguard": "E0302",
    "shadow_warlock": "E0303",
    "shadow_knight": "E0304",
    # 章5
    "eternal_sentinel": "E0401",
    "eternal_chrono": "E0402",
    "eternal_archer": "E0403",
    "eternal_priest": "E0404",
    # 精英
    "elite_necromancer": "E9001",
    "elite_alpha_wolf": "E9002",
    "elite_frost_wyrm": "E9003",
    "elite_ice_lord": "E9004",
    "elite_infernal": "E9005",
    "elite_dark_iron": "E9006",
    "elite_doomguard": "E9007",
    "elite_shadow_priest": "E9008",
    "elite_titan_construct": "E9009",
    "elite_void_walker": "E9010",
    # Boss
    "boss_lich_forest": "E9101",
    "boss_ancient_treant": "E9102",
    "boss_frost_queen": "E9103",
    "boss_frost_lich": "E9104",
    "boss_ragnaros": "E9105",
    "boss_deathwing": "E9106",
    "boss_archimonde": "E9107",
    "boss_kiljaeden": "E9108",
    "boss_titan_watcher": "E9109",
    "boss_eternal_lord": "E9110",
}


# 完整敌人定义（enemy_config.gd 原数据）
# 格式：(legacy_id, name, chapter, hp, dmg, category, combat_type, gold, drop_relic, drop_reroll,
#        phases[(threshold, actions[(type, value, desc, scalable)])],
#        quotes{enter/death/attack/hurt/low_hp: [str]})
# 为了保持紧凑，这里用列表而不是 dict
ENEMIES = [
    # ---- 章1 森林 普通 ----
    ("forest_ghoul", "食尸鬼", 1, 28, 7, "NORMAL", "WARRIOR", 20, False, 0,
     [(0, [("ATTACK", 7, "", True), ("ATTACK", 9, "撕咬", True), ("SKILL", 1, "虚弱", False)])],
     {"enter": ["嘎嘎……新鲜的肉……", "从坟墓里爬出来了……"],
      "death": ["骨头……散了……", "回到……土里……"],
      "attack": ["撕！", "咬碎你！", "嘎嘎嘎！"],
      "hurt": ["嘎！", "腐肉……掉了……"],
      "low_hp": ["不……还没吃饱……"]}),
    ("forest_spider", "剧毒蛛母", 1, 18, 3, "NORMAL", "RANGER", 20, False, 0,
     [(0, [("SKILL", 2, "剧毒", False), ("ATTACK", 4, "", True), ("ATTACK", 4, "", True)])],
     {"enter": ["嘶嘶……陷阱已经布好了……"], "death": ["嘶……蛛卵……会替我……"],
      "attack": ["毒牙！", "吐丝！", "缠住你！"], "hurt": ["嘶！", "我的……腿！"],
      "low_hp": ["蛛巢……不会忘记你……"]}),
    ("forest_treant", "腐化树人", 1, 42, 4, "NORMAL", "GUARDIAN", 20, False, 0,
     [(0, [("DEFEND", 8, "", True), ("ATTACK", 5, "", True), ("DEFEND", 6, "", True), ("ATTACK", 7, "根须缠绕", True)])],
     {"enter": ["这片……森林……不欢迎你……"], "death": ["森林……会记住……"],
      "attack": ["根须！", "大地之力！"], "hurt": ["树皮……裂了……"],
      "low_hp": ["我的根……断了……"]}),
    ("forest_banshee", "哀嚎女妖", 1, 16, 3, "NORMAL", "CASTER", 20, False, 0,
     [(0, [("SKILL", 1, "易伤", False), ("ATTACK", 5, "", True), ("SKILL", 1, "虚弱", False)])],
     {"enter": ["啊啊啊——！"], "death": ["终于……安息了……"],
      "attack": ["尖叫！", "死亡之歌！"], "hurt": ["（刺耳尖啸）"],
      "low_hp": ["最后……一曲……"]}),
    ("forest_wolf_priest", "月光狼灵", 1, 20, 2, "NORMAL", "PRIEST", 20, False, 0,
     [(0, [("SKILL", 2, "剧毒", False), ("SKILL", 1, "易伤", False), ("ATTACK", 4, "", True)])],
     {"enter": ["呜——月光指引着我……"], "death": ["月光……暗了……"],
      "attack": ["狼牙！", "月光之噬！"], "hurt": ["嗷！"],
      "low_hp": ["月光……给我力量……"]}),
    # ---- 章2 冰封 普通 ----
    ("ice_yeti", "雪原雪人", 2, 36, 9, "NORMAL", "WARRIOR", 20, False, 0,
     [(0, [("ATTACK", 9, "", True), ("ATTACK", 11, "冰拳", True)])],
     {"enter": ["吼————！"], "death": ["冰……碎了……"],
      "attack": ["砸！", "冰拳！", "吼！"], "hurt": ["吼！疼！"],
      "low_hp": ["不会……倒下……"]}),
    ("ice_mage", "霜寒女巫", 2, 18, 4, "NORMAL", "CASTER", 20, False, 0,
     [(0, [("SKILL", 1, "冻结", False), ("ATTACK", 6, "", True), ("SKILL", 2, "虚弱", False)])],
     {"enter": ["冰霜……会冻结一切……"], "death": ["冰……碎了……"],
      "attack": ["冰锥！", "寒冰箭！"], "hurt": ["冰盾……裂了……"],
      "low_hp": ["暴风雪……最后的咏唱……"]}),
    ("ice_wolf", "霜鬃狼", 2, 22, 5, "NORMAL", "RANGER", 20, False, 0,
     [(0, [("ATTACK", 5, "", True), ("ATTACK", 7, "冰霜撕咬", True), ("SKILL", 1, "灼烧", False)])],
     {"enter": ["（低沉的咆哮）"], "death": ["呜……"],
      "attack": ["嗷！", "撕咬！", "冰牙！"], "hurt": ["嗷呜！"],
      "low_hp": ["群狼……会替我报仇……"]}),
    ("ice_golem", "寒冰石像", 2, 44, 4, "NORMAL", "GUARDIAN", 20, False, 0,
     [(0, [("DEFEND", 10, "", True), ("ATTACK", 5, "", True), ("DEFEND", 8, "", True)])],
     {"enter": ["（冰晶嘎吱作响）"], "death": ["（碎裂成冰块）"],
      "attack": ["碾压！", "冰拳！"], "hurt": ["裂缝……"],
      "low_hp": ["还能……守住……"]}),
    # ---- 章3 熔岩 普通 ----
    ("lava_hound", "地狱火犬", 3, 30, 8, "NORMAL", "WARRIOR", 20, False, 0,
     [(0, [("ATTACK", 8, "", True), ("ATTACK", 10, "烈焰撕咬", True), ("SKILL", 2, "灼烧", False)])],
     {"enter": ["（烈焰从口中喷出）"], "death": ["火……灭了……"],
      "attack": ["烈焰！", "烧！", "吞噬！"], "hurt": ["（痛苦嚎叫）"],
      "low_hp": ["最后……一口火焰……"]}),
    ("lava_imp", "小恶魔", 3, 16, 4, "NORMAL", "CASTER", 20, False, 0,
     [(0, [("SKILL", 2, "灼烧", False), ("ATTACK", 5, "", True), ("SKILL", 1, "易伤", False), ("ATTACK", 6, "火球", True)])],
     {"enter": ["嘻嘻嘻！又来送死的！"], "death": ["嘻……不好玩了……"],
      "attack": ["接火球！", "嘻嘻！烫吧！"], "hurt": ["哎呀！"],
      "low_hp": ["不行了……要逃了……才怪！"]}),
    ("lava_guardian", "黑铁卫士", 3, 48, 5, "NORMAL", "GUARDIAN", 20, False, 0,
     [(0, [("DEFEND", 12, "", True), ("ATTACK", 6, "", True), ("DEFEND", 8, "", True), ("ATTACK", 8, "锻造重击", True)])],
     {"enter": ["黑铁之盾，坚不可摧！"], "death": ["盾……碎了……"],
      "attack": ["锤击！", "黑铁之力！"], "hurt": ["叮！"],
      "low_hp": ["只要……盾还在……就不会倒！"]}),
    ("lava_shaman", "火焰萨满", 3, 22, 3, "NORMAL", "PRIEST", 20, False, 0,
     [(0, [("SKILL", 2, "灼烧", False), ("SKILL", 1, "力量", False), ("ATTACK", 5, "", True)])],
     {"enter": ["烈焰之灵……降临吧！"], "death": ["火灵……离开了我……"],
      "attack": ["烈焰冲击！", "焚烧！"], "hurt": ["火盾……碎了……"],
      "low_hp": ["最后的祈祷……"]}),
    # ---- 章4 暗影 普通 ----
    ("shadow_assassin", "暗影刺客", 4, 24, 12, "NORMAL", "RANGER", 20, False, 0,
     [(0, [("ATTACK", 12, "背刺", True), ("SKILL", 2, "剧毒", False), ("ATTACK", 8, "", True)])],
     {"enter": ["（从阴影中浮现）"], "death": ["影子……消散了……"],
      "attack": ["背刺！", "影杀！"], "hurt": ["嘶……被发现了……"],
      "low_hp": ["影遁……最后一击……"]}),
    ("shadow_felguard", "邪能卫兵", 4, 46, 6, "NORMAL", "GUARDIAN", 20, False, 0,
     [(0, [("ATTACK", 7, "", True), ("DEFEND", 14, "", True), ("ATTACK", 9, "邪能重斩", True)])],
     {"enter": ["受主人之命……消灭一切入侵者！"], "death": ["主人……恕我……"],
      "attack": ["邪能斩！", "毁灭！"], "hurt": ["邪能护甲……"],
      "low_hp": ["主人的力量……赐予我……"]}),
    ("shadow_warlock", "邪能术士", 4, 20, 5, "NORMAL", "CASTER", 20, False, 0,
     [(0, [("SKILL", 2, "剧毒", False), ("ATTACK", 6, "", True), ("SKILL", 2, "灼烧", False), ("ATTACK", 7, "暗影箭", True)])],
     {"enter": ["邪能……是最强大的力量！"], "death": ["不……我的灵魂……"],
      "attack": ["暗影箭！", "燃烧吧！"], "hurt": ["灵魂石……碎了……"],
      "low_hp": ["生命分流！"]}),
    ("shadow_knight", "堕落死亡骑士", 4, 34, 10, "NORMAL", "WARRIOR", 20, False, 0,
     [(0, [("ATTACK", 10, "", True), ("SKILL", 1, "虚弱", False), ("ATTACK", 12, "凋零打击", True)])],
     {"enter": ["曾经……我也是光明的骑士……"], "death": ["光……我又看到了……光……"],
      "attack": ["凋零！", "黑暗之力！"], "hurt": ["这具身体……已经不怕痛了……"],
      "low_hp": ["即便倒下……黑暗……也不会消失……"]}),
    # ---- 章5 永恒 普通 ----
    ("eternal_sentinel", "光铸哨兵", 5, 40, 8, "NORMAL", "GUARDIAN", 20, False, 0,
     [(0, [("DEFEND", 14, "", True), ("ATTACK", 8, "", True), ("DEFEND", 10, "", True), ("ATTACK", 10, "圣光裁决", True)])],
     {"enter": ["此地……不可侵犯。"], "death": ["任务……失败……"],
      "attack": ["裁决！", "净化！"], "hurt": ["圣光护盾……动摇了……"],
      "low_hp": ["即使倒下……光明……永不熄灭……"]}),
    ("eternal_chrono", "时光龙人", 5, 26, 7, "NORMAL", "CASTER", 20, False, 0,
     [(0, [("SKILL", 2, "虚弱", False), ("ATTACK", 8, "时光冲击", True), ("SKILL", 1, "冻结", False)])],
     {"enter": ["你的时间线……出了偏差……"], "death": ["时间线……修复了……"],
      "attack": ["时光逆转！", "沙漏之力！"], "hurt": ["时间流……紊乱了……"],
      "low_hp": ["最后的沙粒……"]}),
    ("eternal_archer", "星界游侠", 5, 22, 10, "NORMAL", "RANGER", 20, False, 0,
     [(0, [("ATTACK", 10, "", True), ("ATTACK", 12, "星辰之箭", True), ("SKILL", 1, "易伤", False)])],
     {"enter": ["星光……指引我的箭矢……"], "death": ["星辰……暗了……"],
      "attack": ["星箭！", "穿透！"], "hurt": ["嘶……"],
      "low_hp": ["最后一箭……"]}),
    ("eternal_priest", "泰坦祭司", 5, 24, 3, "NORMAL", "PRIEST", 20, False, 0,
     [(0, [("SKILL", 2, "力量", False), ("SKILL", 1, "易伤", False), ("ATTACK", 6, "圣光惩击", True)])],
     {"enter": ["泰坦的意志……不容亵渎。"], "death": ["泰坦……我……回来了……"],
      "attack": ["惩击！", "圣光！"], "hurt": ["信仰……不会动摇……"],
      "low_hp": ["圣光……赐予我……"]}),
    # ---- 精英 ----
    ("elite_necromancer", "亡灵巫师", 1, 85, 8, "ELITE", "CASTER", 50, True, 2,
     [(0.4, [("ATTACK", 14, "亡灵大军", True), ("SKILL", 3, "剧毒", False)]),
      (0, [("ATTACK", 8, "", True), ("SKILL", 2, "虚弱", False), ("DEFEND", 12, "", True)])],
     {"enter": ["死者……听从我的召唤！"], "death": ["我的……亡灵们……"],
      "attack": ["亡灵术！", "腐蚀！", "黑暗吞噬！"], "hurt": ["骨盾……碎了？"],
      "low_hp": ["用我的骸骨……召唤最后的亡灵！"]}),
    ("elite_alpha_wolf", "狼人首领", 1, 100, 11, "ELITE", "WARRIOR", 50, True, 2,
     [(0, [("ATTACK", 11, "", True), ("ATTACK", 14, "狂暴撕咬", True), ("SKILL", 2, "力量", False), ("ATTACK", 9, "", True)])],
     {"enter": ["月光之下……狼群为王！"], "death": ["狼王……倒下了……"],
      "attack": ["撕碎！", "狂暴！", "嗷——！"], "hurt": ["疼痛……让我更愤怒！"],
      "low_hp": ["月光……赐予我……"]}),
    ("elite_frost_wyrm", "霜龙幼崽", 2, 95, 10, "ELITE", "CASTER", 50, True, 2,
     [(0.3, [("ATTACK", 18, "寒冰吐息", True), ("SKILL", 2, "冻结", False)]),
      (0, [("ATTACK", 10, "", True), ("SKILL", 2, "虚弱", False), ("DEFEND", 14, "", True), ("ATTACK", 8, "", True)])],
     {"enter": ["（冰冷的咆哮响彻山谷）"], "death": ["（碎裂成无数冰晶）"],
      "attack": ["冰息！", "冻住吧！"], "hurt": ["龙鳞……裂了？"],
      "low_hp": ["最后的……寒冰吐息……"]}),
    ("elite_ice_lord", "冰霜巨人王", 2, 120, 7, "ELITE", "GUARDIAN", 50, True, 2,
     [(0, [("DEFEND", 20, "", True), ("ATTACK", 8, "", True), ("ATTACK", 14, "冰锤粉碎", True), ("SKILL", 1, "冻结", False)])],
     {"enter": ["渺小的生物……敢闯冰封王座？"], "death": ["冰……不灭……"],
      "attack": ["碾碎！", "冰锤！"], "hurt": ["蚊虫叮咬……"],
      "low_hp": ["冰封王座……不会倒塌！"]}),
    ("elite_infernal", "地狱火", 3, 100, 12, "ELITE", "WARRIOR", 50, True, 2,
     [(0, [("ATTACK", 12, "", True), ("ATTACK", 16, "烈焰冲击", True), ("SKILL", 3, "灼烧", False), ("DEFEND", 10, "", True)])],
     {"enter": ["（从天而降，地面龟裂）"], "death": ["烈焰……熄灭了……"],
      "attack": ["烈焰！", "毁灭！", "焚烧一切！"], "hurt": ["石皮……裂了……"],
      "low_hp": ["最后的爆发……"]}),
    ("elite_dark_iron", "黑铁议员", 3, 90, 9, "ELITE", "CASTER", 50, True, 2,
     [(0.4, [("ATTACK", 16, "熔岩之怒", True), ("SKILL", 1, "诅咒锻造", False)]),
      (0, [("ATTACK", 9, "", True), ("SKILL", 2, "灼烧", False), ("DEFEND", 16, "", True)])],
     {"enter": ["黑铁议会……判你死刑！"], "death": ["议会……散了……"],
      "attack": ["熔岩之怒！", "黑铁审判！"], "hurt": ["黑铁……不碎！"],
      "low_hp": ["启动……自毁程序……"]}),
    ("elite_doomguard", "末日守卫", 4, 110, 11, "ELITE", "WARRIOR", 50, True, 2,
     [(0, [("ATTACK", 11, "", True), ("ATTACK", 16, "末日审判", True), ("SKILL", 2, "易伤", False), ("DEFEND", 14, "", True), ("SKILL", 1, "诅咒注入", False)])],
     {"enter": ["末日……已经降临。"], "death": ["军团……不灭……"],
      "attack": ["末日审判！", "灵魂撕裂！"], "hurt": ["邪能护甲……动摇了？"],
      "low_hp": ["用我的生命……召唤更强大的恶魔！"]}),
    ("elite_shadow_priest", "暗影大主教", 4, 80, 8, "ELITE", "PRIEST", 50, True, 2,
     [(0.3, [("SKILL", 3, "剧毒", False), ("SKILL", 3, "灼烧", False)]),
      (0, [("ATTACK", 8, "", True), ("SKILL", 2, "虚弱", False), ("ATTACK", 10, "精神鞭笞", True), ("SKILL", 2, "剧毒", False)])],
     {"enter": ["暗影的低语……你听到了吗？"], "death": ["暗影……弥散了……"],
      "attack": ["精神鞭笞！", "暗影之触！"], "hurt": ["心灵屏障……裂了……"],
      "low_hp": ["暗影……形态——最终手段！"]}),
    ("elite_titan_construct", "泰坦守护者", 5, 130, 10, "ELITE", "GUARDIAN", 50, True, 2,
     [(0, [("DEFEND", 22, "", True), ("ATTACK", 10, "", True), ("ATTACK", 18, "泰坦之锤", True), ("SKILL", 2, "虚弱", False)])],
     {"enter": ["入侵者检测完毕。启动消灭程序。"], "death": ["系统……崩溃……"],
      "attack": ["泰坦之锤！", "消灭目标！"], "hurt": ["护盾……承受冲击……"],
      "low_hp": ["核心过载……启动自毁倒计时……"]}),
    ("elite_void_walker", "虚空行者", 5, 90, 13, "ELITE", "CASTER", 50, True, 2,
     [(0.35, [("ATTACK", 20, "虚空爆裂", True), ("SKILL", 1, "诅咒注入", False)]),
      (0, [("ATTACK", 13, "", True), ("SKILL", 2, "易伤", False), ("ATTACK", 10, "", True), ("SKILL", 2, "虚弱", False)])],
     {"enter": ["虚空……无处不在……"], "death": ["虚空……会记住你……"],
      "attack": ["虚空爆裂！", "维度撕裂！"], "hurt": ["虚空……波动了……"],
      "low_hp": ["虚空的全部力量……释放！"]}),
    # ---- Boss ----
    ("boss_lich_forest", "枯骨巫妖", 1, 150, 10, "BOSS", "CASTER", 60, True, 0,
     [(0.4, [("ATTACK", 16, "亡灵风暴", True), ("SKILL", 2, "灼烧", False), ("ATTACK", 14, "骸骨之矛", True), ("SKILL", 1, "诅咒", False), ("DEFEND", 15, "", True)]),
      (0, [("ATTACK", 8, "", True), ("ATTACK", 8, "", True), ("SKILL", 2, "虚弱", False), ("SKILL", 1, "易伤", False), ("DEFEND", 15, "", True)])],
     {"enter": ["哈哈哈……又一个活人，送上门来了。"], "death": ["我的……灵魂宝石……不——！"],
      "attack": ["亡灵风暴！", "骸骨之矛！"], "hurt": ["灵魂宝石……动摇了……"],
      "low_hp": ["灵魂宝石……碎裂吧——！"]}),
    ("boss_ancient_treant", "远古树王", 1, 300, 15, "BOSS", "GUARDIAN", 0, False, 0,
     [(0.5, [("ATTACK", 22, "大地之怒", True), ("DEFEND", 30, "", True), ("ATTACK", 18, "", True), ("SKILL", 3, "剧毒", False)]),
      (0, [("DEFEND", 20, "", True), ("ATTACK", 12, "", True), ("SKILL", 2, "虚弱", False), ("ATTACK", 15, "", True)])],
     {"enter": ["千年……未曾有人……走到这里。"], "death": ["你……是第一个……砍倒我的人……"],
      "attack": ["大地之怒！", "根须绞杀！"], "hurt": ["不过是……树皮划痕……"],
      "low_hp": ["大地啊……赐予我……最后的力量——！"]}),
    ("boss_frost_queen", "霜寒女王", 2, 160, 10, "BOSS", "CASTER", 60, True, 0,
     [(0.4, [("ATTACK", 18, "暴风雪", True), ("SKILL", 2, "冻结", False), ("ATTACK", 14, "", True), ("SKILL", 1, "碎裂诅咒", False), ("DEFEND", 16, "", True)]),
      (0, [("ATTACK", 9, "", True), ("SKILL", 1, "冻结", False), ("ATTACK", 9, "", True), ("SKILL", 2, "虚弱", False), ("DEFEND", 14, "", True)])],
     {"enter": ["冰封山脉的女王……亲自迎接你。"], "death": ["（冰雕碎裂）"],
      "attack": ["暴风雪！", "冰封！"], "hurt": ["我的冰甲……裂了？"],
      "low_hp": ["冰封……整个世界吧——！"]}),
    ("boss_frost_lich", "霜之巫妖王", 2, 320, 15, "BOSS", "WARRIOR", 0, False, 0,
     [(0.5, [("ATTACK", 28, "霜之哀伤", True), ("ATTACK", 20, "", True), ("SKILL", 3, "剧毒", False), ("DEFEND", 28, "", True)]),
      (0, [("ATTACK", 14, "", True), ("SKILL", 2, "冻结", False), ("ATTACK", 18, "", True), ("SKILL", 2, "虚弱", False)])],
     {"enter": ["跪下……在巫妖王面前。"], "death": ["永恒的寒冬……终结了？"],
      "attack": ["霜之哀伤！", "臣服于寒冰！"], "hurt": ["不过是……暖风拂面。"],
      "low_hp": ["所有人……都将臣服于寒冰王座——！"]}),
    ("boss_ragnaros", "炎魔之王", 3, 200, 12, "BOSS", "WARRIOR", 60, True, 0,
     [(0.4, [("ATTACK", 20, "岩浆之锤", True), ("SKILL", 3, "灼烧", False), ("ATTACK", 16, "烈焰之手", True), ("DEFEND", 14, "", True)]),
      (0, [("ATTACK", 12, "", True), ("SKILL", 2, "灼烧", False), ("ATTACK", 10, "", True), ("DEFEND", 12, "", True)])],
     {"enter": ["太早了……你唤醒我太早了！"], "death": ["我会……回来的……"],
      "attack": ["岩浆之锤！", "烈焰冲击！"], "hurt": ["渣渣！你敢伤我？"],
      "low_hp": ["烈焰……最后的爆发——！"]}),
    ("boss_deathwing", "熔火死翼", 3, 380, 16, "BOSS", "CASTER", 0, False, 0,
     [(0.5, [("ATTACK", 30, "大灾变", True), ("ATTACK", 22, "", True), ("SKILL", 4, "灼烧", False), ("DEFEND", 30, "", True)]),
      (0, [("SKILL", 3, "灼烧", False), ("ATTACK", 14, "", True), ("SKILL", 2, "易伤", False), ("ATTACK", 20, "熔岩吐息", True)])],
     {"enter": ["大灾变……来临了！"], "death": ["（咆哮着坠入岩浆）"],
      "attack": ["大灾变！", "熔岩吐息！"], "hurt": ["你伤到了……我的钢铁之躯？"],
      "low_hp": ["即使我倒下……世界……也已面目全非——！"]}),
    ("boss_archimonde", "深渊领主", 4, 200, 11, "BOSS", "CASTER", 60, True, 0,
     [(0.4, [("ATTACK", 18, "暗影之手", True), ("SKILL", 2, "灼烧", False), ("ATTACK", 14, "邪能风暴", True), ("SKILL", 1, "诅咒注入", False), ("DEFEND", 16, "", True)]),
      (0, [("ATTACK", 10, "", True), ("SKILL", 2, "虚弱", False), ("ATTACK", 9, "", True), ("SKILL", 2, "剧毒", False), ("DEFEND", 14, "", True)])],
     {"enter": ["燃烧军团……势不可挡！"], "death": ["我会……在扭曲虚空中……重生！"],
      "attack": ["暗影之手！", "邪能风暴！"], "hurt": ["你……竟敢？"],
      "low_hp": ["燃烧吧——！"]}),
    ("boss_kiljaeden", "暗影之王", 4, 380, 16, "BOSS", "CASTER", 0, False, 0,
     [(0.5, [("ATTACK", 28, "黑暗终焉", True), ("ATTACK", 22, "", True), ("SKILL", 3, "剧毒", False), ("DEFEND", 30, "", True)]),
      (0, [("SKILL", 4, "灼烧", False), ("ATTACK", 14, "", True), ("SKILL", 2, "虚弱", False), ("ATTACK", 20, "邪能陨石", True)])],
     {"enter": ["欺骗者……来了。"], "death": ["不可能……欺骗者……怎会被欺骗……"],
      "attack": ["黑暗终焉！", "邪能陨石！"], "hurt": ["有趣……你确实……有些能耐。"],
      "low_hp": ["用虚空的全部力量——毁灭这个世界！"]}),
    ("boss_titan_watcher", "泰坦看守者", 5, 200, 12, "BOSS", "GUARDIAN", 60, True, 0,
     [(0.4, [("ATTACK", 18, "泰坦审判", True), ("DEFEND", 22, "", True), ("ATTACK", 16, "秩序之光", True), ("SKILL", 2, "易伤", False)]),
      (0, [("DEFEND", 18, "", True), ("ATTACK", 10, "", True), ("SKILL", 2, "虚弱", False), ("ATTACK", 12, "", True)])],
     {"enter": ["泰坦的秩序……不容亵渎。"], "death": ["秩序……被打破了……"],
      "attack": ["泰坦审判！", "秩序之光！"], "hurt": ["损伤……在可控范围内……"],
      "low_hp": ["启动……最终审判协议——！"]}),
    ("boss_eternal_lord", "永恒主宰", 5, 480, 18, "BOSS", "CASTER", 0, False, 0,
     [(0.5, [("ATTACK", 28, "终极之光", True), ("ATTACK", 22, "", True), ("SKILL", 3, "剧毒", False), ("DEFEND", 30, "", True)]),
      (0, [("SKILL", 4, "灼烧", False), ("ATTACK", 14, "", True), ("SKILL", 2, "虚弱", False), ("ATTACK", 20, "", True)])],
     {"enter": ["永恒……在此。渺小的骰子掷者，你的终点……就是今天。"],
      "death": ["你……究竟……是什么？"],
      "attack": ["终极之光！", "永恒之力，碾碎你！"],
      "hurt": ["哼……有点意思。"],
      "low_hp": ["永恒……动摇了……但我绝不会……就此终结！终极之光——爆发！"]}),
]


def build_enemy():
    wb = Workbook()

    # ===== Sheet 1: base =====
    ws = wb.active
    ws.title = "base"
    write_header(ws, [
        ("id", "string", "敌人编号 E+4位。章节段号：1章0001-/2章0101-/3章0201-/4章0301-/5章0401-；精英E9001-；Boss E9101-"),
        ("legacy_key", "string", "代码旧 key（forest_ghoul 等）"),
        ("name", "string", "中文名"),
        ("chapter", "int", "所属章节 1-5"),
        ("category", "string", "类型：NORMAL/ELITE/BOSS"),
        ("combat_type", "string", "战斗类型：WARRIOR/GUARDIAN/RANGER/CASTER/PRIEST"),
        ("base_hp", "int", "基础血量（会被章节/深度系数缩放）"),
        ("base_dmg", "int", "基础攻击力"),
        ("drop_gold", "int", "击杀金币"),
        ("drop_relic", "bool", "是否掉落遗物 1是0否"),
        ("drop_reroll_reward", "int", "掉落重投奖励次数"),
        ("phase_group", "ref", "阶段组外键，指向 phases sheet 的 phase_group"),
        ("quote_group", "ref", "台词组外键，指向 quotes sheet 的 quote_group"),
    ], widths=[8, 24, 16, 8, 10, 12, 10, 10, 10, 10, 14, 14, 14])

    for r, (legacy, name, chapter, hp, dmg, cat, ct, gold, drop_r, drop_rr, phases, quotes) in enumerate(ENEMIES, start=4):
        eid = ENEMY_ID_MAP[legacy]
        pg = "PG" + eid[1:]  # phase_group 用 PG + 敌人号段
        qg = "QG" + eid[1:]
        ws.cell(row=r, column=1, value=eid)
        ws.cell(row=r, column=2, value=legacy)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=chapter)
        ws.cell(row=r, column=5, value=cat)
        ws.cell(row=r, column=6, value=ct)
        ws.cell(row=r, column=7, value=hp)
        ws.cell(row=r, column=8, value=dmg)
        ws.cell(row=r, column=9, value=gold)
        ws.cell(row=r, column=10, value=1 if drop_r else 0)
        ws.cell(row=r, column=11, value=drop_rr)
        ws.cell(row=r, column=12, value=pg)
        ws.cell(row=r, column=13, value=qg)

    # ===== Sheet 2: phases =====
    ws_ph = wb.create_sheet("phases")
    write_header(ws_ph, [
        ("phase_group", "ref", "阶段组编号 PG+4位"),
        ("phase_idx", "int", "阶段序号 0 起"),
        ("hp_threshold", "float", "HP 百分比阈值（低于此值切换。0 表示无条件首个阶段）"),
        ("action_group", "ref", "该阶段的行动组，指向 actions sheet"),
    ], widths=[12, 10, 14, 14])

    # ===== Sheet 3: actions =====
    ws_ac = wb.create_sheet("actions")
    write_header(ws_ac, [
        ("action_group", "ref", "行动组编号 AG+6位（AG+敌人号段+阶段序号）"),
        ("action_idx", "int", "行动顺序 0 起"),
        ("type", "string", "类型：ATTACK/DEFEND/SKILL"),
        ("base_value", "int", "基础数值（攻击=伤害 / 防御=护甲 / 技能=数值）"),
        ("description", "string", "行动描述（SKILL 类填状态名，如 灼烧/虚弱/剧毒）"),
        ("scalable", "bool", "是否受章节/深度缩放 1是0否（SKILL 类通常0）"),
    ], widths=[14, 10, 10, 12, 18, 10])

    phase_rows = []
    action_rows = []
    for legacy, name, chapter, hp, dmg, cat, ct, gold, drop_r, drop_rr, phases, quotes in ENEMIES:
        eid = ENEMY_ID_MAP[legacy]
        pg = "PG" + eid[1:]
        for p_idx, (threshold, actions) in enumerate(phases):
            ag = f"AG{eid[1:]}{p_idx}"  # AG + 敌人号段 + 阶段序号 = AG00010 / AG91010
            phase_rows.append((pg, p_idx, threshold, ag))
            for a_idx, (a_type, a_value, a_desc, a_scalable) in enumerate(actions):
                action_rows.append((ag, a_idx, a_type, a_value, a_desc, 1 if a_scalable else 0))

    for r, data in enumerate(phase_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws_ph.cell(row=r, column=col, value=v)
    for r, data in enumerate(action_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws_ac.cell(row=r, column=col, value=v)

    # ===== Sheet 4: quotes =====
    ws_q = wb.create_sheet("quotes")
    write_header(ws_q, [
        ("quote_group", "ref", "台词组编号 QG+4位"),
        ("event", "string", "触发事件：enter/death/attack/hurt/low_hp"),
        ("text", "string", "台词文本"),
    ], widths=[12, 10, 60])

    quote_rows = []
    for legacy, name, chapter, hp, dmg, cat, ct, gold, drop_r, drop_rr, phases, quotes in ENEMIES:
        eid = ENEMY_ID_MAP[legacy]
        qg = "QG" + eid[1:]
        for event in ("enter", "death", "attack", "hurt", "low_hp"):
            for text in quotes.get(event, []):
                quote_rows.append((qg, event, text))

    for r, data in enumerate(quote_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws_q.cell(row=r, column=col, value=v)

    out_path = os.path.join(EXCEL_DIR, "enemy.xlsx")
    wb.save(out_path)
    print(f"[OK] {out_path} · base={len(ENEMIES)} phases={len(phase_rows)} actions={len(action_rows)} quotes={len(quote_rows)}")


if __name__ == "__main__":
    build_enemy()
