"""
一键生成全部 Excel 配置表。
会顺序调用 build_excel_enums / balance / class / dice / relic / enemy，
并把所有实际 ID 合并写入 _enums.xlsx 的 id_map sheet。

运行：python config/tools/build_excel_all.py
"""
import os
import subprocess
import sys
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
EXCEL_DIR = os.path.join(PROJECT_ROOT, "config", "excel")


def run(script):
    """运行同目录下的子脚本，直接继承父进程 stdout（避免 Windows 编码问题）"""
    path = os.path.join(SCRIPT_DIR, script)
    result = subprocess.run([sys.executable, path])
    if result.returncode != 0:
        print(f"[FAIL] {script}")
        sys.exit(1)


def merge_id_map():
    """把 class / dice / relic / enemy 的 id 合并写回 _enums.xlsx id_map"""
    enums_path = os.path.join(EXCEL_DIR, "_enums.xlsx")
    wb = load_workbook(enums_path)
    ws = wb["id_map"]

    # 先找到已有数据的最后一行
    last_row = ws.max_row
    # 数据从第 4 行开始
    if last_row < 4:
        last_row = 3

    extra_rows = []

    # 骰子
    dice_wb = load_workbook(os.path.join(EXCEL_DIR, "dice.xlsx"))
    dice_ws = dice_wb["base"]
    for r in range(4, dice_ws.max_row + 1):
        did = dice_ws.cell(row=r, column=1).value
        legacy = dice_ws.cell(row=r, column=2).value
        name = dice_ws.cell(row=r, column=4).value
        if did:
            extra_rows.append((did, "骰子", name, legacy, ""))

    # 遗物
    relic_wb = load_workbook(os.path.join(EXCEL_DIR, "relic.xlsx"))
    relic_ws = relic_wb["base"]
    for r in range(4, relic_ws.max_row + 1):
        rid = relic_ws.cell(row=r, column=1).value
        legacy = relic_ws.cell(row=r, column=2).value
        name = relic_ws.cell(row=r, column=3).value
        if rid:
            extra_rows.append((rid, "遗物", name, legacy, ""))

    # 敌人
    enemy_wb = load_workbook(os.path.join(EXCEL_DIR, "enemy.xlsx"))
    enemy_ws = enemy_wb["base"]
    for r in range(4, enemy_ws.max_row + 1):
        eid = enemy_ws.cell(row=r, column=1).value
        legacy = enemy_ws.cell(row=r, column=2).value
        name = enemy_ws.cell(row=r, column=3).value
        if eid:
            extra_rows.append((eid, "敌人", name, legacy, ""))

    # 写入
    start_row = last_row + 1
    for i, data in enumerate(extra_rows):
        for col, v in enumerate(data, start=1):
            ws.cell(row=start_row + i, column=col, value=v)

    wb.save(enums_path)
    print(f"[OK] id_map merged · +{len(extra_rows)} rows")


def main():
    print("==== 生成所有 Excel 配置表 ====")
    run("build_excel_enums.py")
    run("build_excel_balance.py")
    run("build_excel_class.py")
    run("build_excel_dice.py")
    run("build_excel_relic.py")
    run("build_excel_enemy.py")
    merge_id_map()
    print("==== DONE ====")


if __name__ == "__main__":
    main()
