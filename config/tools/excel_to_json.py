"""
Excel → JSON 转换器
每次策划改完 Excel 后运行这个脚本，生成 config/json/ 下的 json 文件供 Godot 读取。

运行：python config/tools/excel_to_json.py
"""
import os
import json
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
EXCEL_DIR = os.path.join(PROJECT_ROOT, "config", "excel")
JSON_DIR = os.path.join(PROJECT_ROOT, "config", "json")
os.makedirs(JSON_DIR, exist_ok=True)


def parse_value(raw, type_name):
    """按字段类型解析 Excel 值。空/None 按类型返回默认"""
    if raw is None or raw == "":
        if type_name == "int":
            return 0
        if type_name == "float":
            return 0.0
        if type_name == "bool":
            return False
        if type_name in ("string", "ref", "enum"):
            return ""
        if type_name in ("intarray", "floatarray", "strarray"):
            return []
        if type_name == "json":
            return None
        return None

    t = type_name.strip().lower()
    if t == "int":
        return int(raw)
    if t == "float":
        return float(raw)
    if t == "bool":
        if isinstance(raw, bool):
            return raw
        if isinstance(raw, (int, float)):
            return bool(raw)
        s = str(raw).strip().lower()
        return s in ("1", "true", "yes", "y", "t")
    if t == "string" or t == "ref" or t == "enum":
        return str(raw)
    if t == "intarray":
        return [int(x.strip()) for x in str(raw).split(",") if x.strip() != ""]
    if t == "floatarray":
        return [float(x.strip()) for x in str(raw).split(",") if x.strip() != ""]
    if t == "strarray":
        return [x.strip() for x in str(raw).split(";") if x.strip() != ""]
    if t == "json":
        if isinstance(raw, (dict, list)):
            return raw
        return json.loads(str(raw))
    return raw


def read_sheet(ws):
    """读取一个 sheet 为 list[dict]，跳过 1-3 行表头"""
    # 第 1 行字段名，第 2 行类型，第 3 行备注
    field_names = []
    field_types = []
    col = 1
    while True:
        name = ws.cell(row=1, column=col).value
        if name is None or name == "":
            break
        field_names.append(str(name).strip())
        tval = ws.cell(row=2, column=col).value or "string"
        field_types.append(str(tval).strip())
        col += 1

    records = []
    for r in range(4, ws.max_row + 1):
        # 如果本行所有字段都空，跳过
        row_vals = [ws.cell(row=r, column=c + 1).value for c in range(len(field_names))]
        if all(v is None or v == "" for v in row_vals):
            continue
        rec = {}
        for i, name in enumerate(field_names):
            rec[name] = parse_value(row_vals[i], field_types[i])
        records.append(rec)
    return records


def convert(xlsx_name, out_json_name=None, sheet_mapping=None):
    """转换一个 xlsx。
    sheet_mapping: None → 默认把每个 sheet 作为顶层 key；
                   否则是 dict[sheet_name -> json_key]
    """
    in_path = os.path.join(EXCEL_DIR, xlsx_name)
    if not os.path.exists(in_path):
        print(f"[SKIP] {xlsx_name} not found")
        return
    wb = load_workbook(in_path, data_only=True)

    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        key = sheet_mapping.get(sheet_name, sheet_name) if sheet_mapping else sheet_name
        result[key] = read_sheet(ws)

    out_name = out_json_name or (os.path.splitext(xlsx_name)[0] + ".json")
    out_path = os.path.join(JSON_DIR, out_name)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"[OK] {out_path}")


def convert_balance():
    """balance 比较特殊，value 列要按 type 再二次解析（Excel 里所有 value 都是字符串存储）"""
    in_path = os.path.join(EXCEL_DIR, "balance.xlsx")
    wb = load_workbook(in_path, data_only=True)
    ws = wb["balance"]

    result = {}
    for r in range(4, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        raw = ws.cell(row=r, column=2).value
        type_ = ws.cell(row=r, column=3).value or "string"
        if key is None or key == "":
            continue
        result[str(key)] = parse_value(raw, type_)

    out_path = os.path.join(JSON_DIR, "balance.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"[OK] {out_path} · {len(result)} keys")


def main():
    print("==== Excel → JSON 转换 ====")
    convert("_enums.xlsx", "enums.json")
    convert_balance()
    convert("class.xlsx")
    convert("dice.xlsx")
    convert("relic.xlsx")
    convert("enemy.xlsx")
    print("==== DONE ====")


if __name__ == "__main__":
    main()
