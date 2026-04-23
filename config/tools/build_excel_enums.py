"""
一次性生成 Excel 配置表：_enums.xlsx
包含全局 ID 映射、效果类型枚举、状态类型枚举、触发时机枚举、稀有度、元素等

运行：python config/tools/build_excel_enums.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
EXCEL_DIR = os.path.join(PROJECT_ROOT, "config", "excel")
os.makedirs(EXCEL_DIR, exist_ok=True)

# 样式常量
HEADER_FILL_NAME = PatternFill(start_color="FFB6D7A8", end_color="FFB6D7A8", fill_type="solid")
HEADER_FILL_TYPE = PatternFill(start_color="FFFFE599", end_color="FFFFE599", fill_type="solid")
HEADER_FILL_DESC = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")
HEADER_FONT = Font(bold=True)


def write_header(ws, headers):
    """写 3 行表头
    headers: list of (字段名, 类型, 描述) 三元组
    """
    for col, (name, type_, desc) in enumerate(headers, start=1):
        c1 = ws.cell(row=1, column=col, value=name)
        c1.fill = HEADER_FILL_NAME
        c1.font = HEADER_FONT
        c1.alignment = Alignment(horizontal="center")

        c2 = ws.cell(row=2, column=col, value=type_)
        c2.fill = HEADER_FILL_TYPE
        c2.alignment = Alignment(horizontal="center")

        c3 = ws.cell(row=3, column=col, value=desc)
        c3.fill = HEADER_FILL_DESC
        c3.alignment = Alignment(wrap_text=True, vertical="top")

    # 冻结表头
    ws.freeze_panes = "A4"
    # 自动列宽（粗糙）
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    # 行高
    ws.row_dimensions[3].height = 60


def build_enums():
    wb = Workbook()
    # 默认 sheet 命名为 id_map
    ws_map = wb.active
    ws_map.title = "id_map"

    # ===== Sheet 1: id_map 全局 ID 映射 =====
    write_header(ws_map, [
        ("id", "string", "编号，见编号规则（D骰子/R遗物/E敌人/C职业/EG效果组/ET效果类型/ST状态/TR触发/RA稀有度/EL元素）"),
        ("type", "string", "类型：骰子/遗物/敌人/职业/效果组/效果类型/状态/触发/稀有度/元素"),
        ("name", "string", "中文名称"),
        ("legacy_key", "string", "迁移前的旧 key（代码中的 id 字段），保留以便追溯"),
        ("note", "string", "备注"),
    ])
    # 这张表仅作查号用，主数据各自表里写，数据后面由 build_excel_dice 等脚本补充。
    # 这里先放枚举类的 id
    rows = []
    # 稀有度
    rows += [
        ("RA1", "稀有度", "普通", "COMMON", ""),
        ("RA2", "稀有度", "不凡", "UNCOMMON", ""),
        ("RA3", "稀有度", "稀有", "RARE", ""),
        ("RA4", "稀有度", "传说", "LEGENDARY", ""),
        ("RA5", "稀有度", "诅咒", "CURSE", "仅骰子有此级"),
    ]
    # 元素
    rows += [
        ("EL0", "元素", "无", "NORMAL", ""),
        ("EL1", "元素", "火", "FIRE", ""),
        ("EL2", "元素", "冰", "ICE", ""),
        ("EL3", "元素", "雷", "THUNDER", ""),
        ("EL4", "元素", "毒", "POISON", ""),
        ("EL5", "元素", "圣", "HOLY", ""),
        ("EL6", "元素", "暗", "SHADOW", ""),
    ]
    # 职业
    rows += [
        ("C00", "职业", "通用", "", "表示非职业专属，所有职业可用"),
        ("C01", "职业", "嗜血狂战", "warrior", ""),
        ("C02", "职业", "星界魔导", "mage", ""),
        ("C03", "职业", "影锋刺客", "rogue", ""),
    ]
    for r, data in enumerate(rows, start=4):
        for col, v in enumerate(data, start=1):
            ws_map.cell(row=r, column=col, value=v)

    # ===== Sheet 2: enum_effect_type 效果类型枚举 =====
    ws_et = wb.create_sheet("enum_effect_type")
    write_header(ws_et, [
        ("code", "string", "效果类型编号，ET+2位"),
        ("key", "string", "代码中的语义 key"),
        ("desc", "string", "含义说明"),
    ])
    et_rows = [
        ("ET01", "damage", "造成伤害（加法）对应 DiceDef.bonus_damage / RelicDef.damage"),
        ("ET02", "damage_mult", "伤害倍率（乘法）对应 DiceDef.bonus_mult / RelicDef.multiplier"),
        ("ET03", "armor", "获得护甲 对应 DiceDef.armor / RelicDef.armor"),
        ("ET04", "heal", "恢复生命 对应 DiceDef.heal / RelicDef.heal"),
        ("ET05", "pierce", "穿甲 对应 DiceDef.pierce / RelicDef.pierce"),
        ("ET06", "self_damage", "自残 对应 DiceDef.self_damage"),
        ("ET07", "status_to_enemy", "给敌人施加状态，param_ref 用 ST紧凑编码"),
        ("ET08", "status_to_self", "给自身施加状态，param_ref 用 ST紧凑编码"),
        ("ET09", "extra_dice", "衍生骰子，param_key 常用：grant_shadow_die / clone_self"),
        ("ET10", "bonus_on_keep", "保留到下回合的加成 对应 DiceDef.bonus_on_keep / bonus_per_turn_kept"),
        ("ET11", "scale_flag", "缩放标记（条件触发类）param_key 指具体字段名"),
        ("ET12", "special", "特殊逻辑（代码写死）param_key = 字段名，param_value = 值"),
        ("ET20", "gold_bonus", "金币加成 对应 RelicDef.gold_bonus"),
        ("ET21", "draw_bonus", "抽牌加成 对应 RelicDef.draw_count_bonus"),
        ("ET22", "shop_discount", "商店折扣 对应 RelicDef.shop_discount"),
        ("ET23", "free_rerolls", "免费重投 对应 RelicDef.free_rerolls"),
        ("ET24", "extra_play", "额外出牌 对应 RelicDef.extra_play"),
        ("ET25", "extra_reroll", "额外重投 对应 RelicDef.extra_reroll"),
        ("ET26", "extra_draw", "额外抽牌 对应 RelicDef.extra_draw"),
    ]
    for r, data in enumerate(et_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws_et.cell(row=r, column=col, value=v)

    # ===== Sheet 3: enum_status_type 状态类型 =====
    ws_st = wb.create_sheet("enum_status_type")
    write_header(ws_st, [
        ("code", "string", "状态编号 ST+2位"),
        ("key", "string", "GameTypes.StatusType 对应枚举名"),
        ("name_cn", "string", "中文名"),
        ("desc", "string", "说明"),
    ])
    st_rows = [
        ("ST01", "POISON", "中毒", "每回合扣等层数血量"),
        ("ST02", "BURN", "灼烧", "每回合扣等层数血量（可与 POISON 叠加不同机制）"),
        ("ST03", "DODGE", "闪避", "下次受伤豁免"),
        ("ST04", "VULNERABLE", "易伤", "受到伤害 +50%"),
        ("ST05", "STRENGTH", "力量", "造成伤害 +等层数"),
        ("ST06", "WEAK", "虚弱", "造成伤害 -25%"),
        ("ST07", "ARMOR", "护甲", "抵挡伤害"),
        ("ST08", "SLOW", "缓慢", "部分行动延迟"),
        ("ST09", "FREEZE", "冻结", "跳过回合"),
    ]
    for r, data in enumerate(st_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws_st.cell(row=r, column=col, value=v)

    # ===== Sheet 4: enum_trigger 遗物触发时机 =====
    ws_tr = wb.create_sheet("enum_trigger")
    write_header(ws_tr, [
        ("code", "string", "触发编号 TR+2位"),
        ("key", "string", "GameTypes.RelicTrigger 对应枚举名"),
        ("desc", "string", "说明"),
    ])
    tr_rows = [
        ("TR01", "ON_PLAY", "每次出牌触发"),
        ("TR02", "ON_KILL", "击杀敌人触发"),
        ("TR03", "ON_REROLL", "重投骰子触发"),
        ("TR04", "ON_TURN_START", "回合开始"),
        ("TR05", "ON_TURN_END", "回合结束"),
        ("TR06", "ON_BATTLE_START", "战斗开始"),
        ("TR07", "ON_BATTLE_END", "战斗结束"),
        ("TR08", "ON_DAMAGE_TAKEN", "受伤触发"),
        ("TR09", "ON_FATAL", "致命伤害时触发（免死类）"),
        ("TR10", "ON_FLOOR_CLEAR", "通关一层触发"),
        ("TR11", "ON_MOVE", "地图移动触发"),
        ("TR99", "PASSIVE", "被动（常驻）"),
    ]
    for r, data in enumerate(tr_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws_tr.cell(row=r, column=col, value=v)

    # ===== Sheet 5: enum_hand_type 牌型（参考用）=====
    ws_ht = wb.create_sheet("enum_hand_type")
    write_header(ws_ht, [
        ("idx", "int", "GameTypes.HandType 枚举序号"),
        ("key", "string", "枚举名"),
        ("name_cn", "string", "中文"),
    ])
    ht_rows = [
        (0, "NORMAL_ATTACK", "普通攻击"),
        (1, "PAIR", "对子"),
        (2, "DOUBLE_PAIR", "连对"),
        (3, "TRIPLE_PAIR", "三连对"),
        (4, "TRIPLET", "三条"),
        (5, "STRAIGHT_3", "顺子3"),
        (6, "STRAIGHT_4", "顺子4"),
        (7, "STRAIGHT_5", "顺子5"),
        (8, "STRAIGHT_6", "顺子6"),
        (9, "SAME_ELEMENT", "同元素"),
        (10, "FULL_HOUSE", "葫芦"),
        (11, "FOUR_OF_KIND", "四条"),
        (12, "FIVE_OF_KIND", "五条"),
        (13, "SIX_OF_KIND", "六条"),
        (14, "ELEMENT_STRAIGHT", "元素顺"),
        (15, "ELEMENT_FULL_HOUSE", "元素葫芦"),
        (16, "ROYAL_ELEMENT_STRAIGHT", "皇家元素顺"),
    ]
    for r, data in enumerate(ht_rows, start=4):
        for col, v in enumerate(data, start=1):
            ws_ht.cell(row=r, column=col, value=v)

    out_path = os.path.join(EXCEL_DIR, "_enums.xlsx")
    wb.save(out_path)
    print(f"[OK] {out_path}")


if __name__ == "__main__":
    build_enums()
