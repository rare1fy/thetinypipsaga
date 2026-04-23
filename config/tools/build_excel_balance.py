"""
一次性生成 balance.xlsx —— 游戏数值常量表
扁平 key + value + type + group 结构，对应 data/game_balance.gd

运行：python config/tools/build_excel_balance.py
"""
import os
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
EXCEL_DIR = os.path.join(PROJECT_ROOT, "config", "excel")

HEADER_FILL_NAME = PatternFill(start_color="FFB6D7A8", end_color="FFB6D7A8", fill_type="solid")
HEADER_FILL_TYPE = PatternFill(start_color="FFFFE599", end_color="FFFFE599", fill_type="solid")
HEADER_FILL_DESC = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")


def write_header(ws, headers):
    for col, (name, type_, desc) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = HEADER_FILL_NAME
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=2, column=col, value=type_)
        c.fill = HEADER_FILL_TYPE
        c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=3, column=col, value=desc)
        c.fill = HEADER_FILL_DESC
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A4"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28
    ws.row_dimensions[3].height = 50


def build_balance():
    wb = Workbook()
    ws = wb.active
    ws.title = "balance"

    write_header(ws, [
        ("key", "string", "扁平键，点号分组：player / status / fury / soul_crystal / shop / campfire / loot / map / chapter / elite / animation / dice_reward"),
        ("value", "string", "值本体。int/float/bool/string直接写数字或文本，数组/对象写 JSON 字符串"),
        ("type", "string", "值类型：int/float/bool/string/intarray/floatarray/json"),
        ("group", "string", "分组名（策划分类用，不进代码）"),
        ("note", "string", "说明 / 调参经验"),
    ])

    # ===== 数据对照 data/game_balance.gd =====
    rows = []

    # --- PLAYER_INITIAL ---
    rows += [
        ("player.hp", 100, "int", "player", "玩家初始血量"),
        ("player.max_hp", 100, "int", "player", "玩家最大血量上限"),
        ("player.armor", 0, "int", "player", "玩家初始护甲"),
        ("player.free_rerolls_per_turn", 1, "int", "player", "每回合免费重投次数"),
        ("player.plays_per_turn", 1, "int", "player", "每回合出牌次数"),
        ("player.souls", 0, "int", "player", "魂晶初始值"),
        ("player.relic_slots", 5, "int", "player", "遗物槽数"),
        ("player.draw_count", 3, "int", "player", "初始抽骰数"),
        ("player.max_draw_count", 6, "int", "player", "抽骰数上限"),
    ]

    # --- DEPTH_SCALING ---
    depth_scaling = [
        {"hpMult": 0.90, "dmgMult": 0.40},
        {"hpMult": 1.10, "dmgMult": 0.50},
        {"hpMult": 1.25, "dmgMult": 0.60},
        {"hpMult": 1.50, "dmgMult": 0.75},
        {"hpMult": 1.20, "dmgMult": 0.65},
        {"hpMult": 1.40, "dmgMult": 0.80},
        {"hpMult": 1.20, "dmgMult": 0.70},
        {"hpMult": 1.80, "dmgMult": 1.00},
        {"hpMult": 1.10, "dmgMult": 0.60},
        {"hpMult": 1.40, "dmgMult": 0.80},
        {"hpMult": 1.60, "dmgMult": 0.90},
        {"hpMult": 1.80, "dmgMult": 1.00},
        {"hpMult": 2.00, "dmgMult": 1.10},
        {"hpMult": 1.30, "dmgMult": 0.80},
        {"hpMult": 2.50, "dmgMult": 1.30},
    ]
    rows.append(("difficulty.depth_scaling", json.dumps(depth_scaling, ensure_ascii=False),
                 "json", "difficulty", "15层难度缩放，每层对 HP 和伤害乘数。-1层默认 hp=0.9,dmg=0.8"))

    # --- STATUS_EFFECT_MULT ---
    rows += [
        ("status.weak_mult", 0.75, "float", "status", "虚弱倍率（造成伤害 ×此值）"),
        ("status.vulnerable_mult", 1.5, "float", "status", "易伤倍率（受到伤害 ×此值）"),
    ]

    # --- FURY_CONFIG 战士血怒 ---
    rows += [
        ("fury.damage_per_stack", 0.15, "float", "fury", "战士血怒每层增加的伤害倍率"),
        ("fury.max_stack", 5, "int", "fury", "战士血怒最大层数"),
        ("fury.armor_at_cap", 5, "int", "fury", "血怒叠满后卖血获得的护甲"),
    ]

    # --- SOUL_CRYSTAL_CONFIG 魂晶 ---
    rows += [
        ("soul_crystal.base_mult", 1.0, "float", "soul_crystal", "魂晶基础倍率"),
        ("soul_crystal.mult_per_depth", 0.2, "float", "soul_crystal", "每层深度的倍率叠加"),
        ("soul_crystal.conversion_rate", 0.15, "float", "soul_crystal", "溢出伤害转化为魂晶的比例"),
    ]

    # --- SHOP_CONFIG ---
    rows += [
        ("shop.relic_count", 3, "int", "shop", "商店遗物数量"),
        ("shop.price_min", 20, "int", "shop", "商店价格下限"),
        ("shop.price_max", 80, "int", "shop", "商店价格上限"),
        ("shop.remove_dice_price", 30, "int", "shop", "移除骰子费用"),
    ]

    # --- CAMPFIRE_CONFIG ---
    rows += [
        ("campfire.rest_heal", 40, "int", "campfire", "营火休息回复血量"),
        ("campfire.upgrade_cost_per_level", 20, "int", "campfire", "遗物升级每级费用"),
        ("campfire.max_relic_level", 5, "int", "campfire", "遗物最大升级等级"),
    ]

    # --- LOOT_CONFIG ---
    rows += [
        ("loot.normal_drop_gold", 25, "int", "loot", "普通敌人掉落金币"),
        ("loot.elite_drop_gold", 50, "int", "loot", "精英敌人掉落金币"),
        ("loot.boss_drop_gold", 80, "int", "loot", "Boss 掉落金币"),
        ("loot.relic_choice_count", 3, "int", "loot", "遗物选择数量"),
    ]

    # --- MAP_CONFIG ---
    rows += [
        ("map.total_layers", 15, "int", "map", "地图总层数"),
        ("map.mid_boss_layer", 7, "int", "map", "中间 Boss 所在层"),
        ("map.rest_before_boss_layers", json.dumps([6, 13]), "json", "map", "Boss 前营火层（JSON 数组）"),
    ]

    # --- CHAPTER_CONFIG ---
    rows += [
        ("chapter.total_chapters", 5, "int", "chapter", "总章节数"),
        ("chapter.names", json.dumps(["幽暗森林", "冰封山脉", "熔岩深渊", "暗影要塞", "永恒之巅"], ensure_ascii=False),
         "json", "chapter", "5 章名称（JSON 字符串数组）"),
        ("chapter.heal_percent", 0.6, "float", "chapter", "章节结束回复百分比"),
        ("chapter.bonus_gold", 75, "int", "chapter", "章节结束奖励金币"),
    ]
    chapter_scaling = [
        {"hpMult": 1.0, "dmgMult": 1.0},
        {"hpMult": 1.25, "dmgMult": 1.15},
        {"hpMult": 1.55, "dmgMult": 1.30},
        {"hpMult": 1.90, "dmgMult": 1.50},
        {"hpMult": 2.30, "dmgMult": 1.70},
    ]
    rows.append(("chapter.scaling", json.dumps(chapter_scaling, ensure_ascii=False),
                 "json", "chapter", "5 章难度缩放（JSON 对象数组）"))

    # --- ELITE_CONFIG ---
    rows += [
        ("elite.hp_threshold", 80, "int", "elite", "精英 HP 阈值"),
        ("elite.boss_hp_threshold", 200, "int", "elite", "Boss HP 阈值"),
        ("elite.boss_curse_hp_ratio", 0.4, "float", "elite", "Boss 诅咒触发血量比"),
        ("elite.armor_mult", 1.5, "float", "elite", "精英护甲乘数"),
        ("elite.boss_armor_mult", 2.0, "float", "elite", "Boss 护甲乘数"),
        ("elite.elite_dice_cycle", 3, "int", "elite", "精英掉骰周期"),
        ("elite.boss_curse_cycle", 2, "int", "elite", "Boss 诅咒周期"),
        ("elite.boss_cracked_dice_cycle", 3, "int", "elite", "Boss 碎裂骰周期"),
        ("elite.elite_armor_cycle", 3, "int", "elite", "精英护甲周期"),
        ("elite.boss_armor_cycle", 2, "int", "elite", "Boss 护甲周期"),
    ]

    # --- ANIMATION_TIMING ---
    rows += [
        ("animation.enemy_death_duration", 1800, "int", "animation", "敌人死亡动画时长（毫秒）"),
        ("animation.enemy_death_cleanup_delay", 2200, "int", "animation", "敌人死亡清理延迟"),
        ("animation.wave_transition_death_buffer", 400, "int", "animation", "波次过渡缓冲"),
        ("animation.boss_entrance_duration", 1200, "int", "animation", "Boss 登场时长"),
        ("animation.attack_effect_duration", 400, "int", "animation", "攻击特效时长"),
        ("animation.victory_enemy_cleanup_delay", 2200, "int", "animation", "胜利敌人清理延迟"),
    ]

    # --- DICE_REWARD_REFRESH ---
    rows += [
        ("dice_reward.base_price", 5, "int", "dice_reward", "刷新骰子奖励基础价格"),
        ("dice_reward.price_multiplier", 2, "int", "dice_reward", "价格倍率（每次刷新翻倍）"),
        ("dice_reward.first_free", True, "bool", "dice_reward", "首次刷新免费"),
    ]

    # 写入
    for r, data in enumerate(rows, start=4):
        for col, v in enumerate(data, start=1):
            # bool 转 1/0 字符串方便 Excel 展示统一
            if isinstance(v, bool):
                v = 1 if v else 0
            ws.cell(row=r, column=col, value=v)

    out_path = os.path.join(EXCEL_DIR, "balance.xlsx")
    wb.save(out_path)
    print(f"[OK] {out_path} · {len(rows)} rows")


if __name__ == "__main__":
    build_balance()
